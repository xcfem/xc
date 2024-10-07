# -*- coding: utf-8 -*-
from __future__ import division
from __future__ import print_function

'''Import/export load combinations.'''

__author__= "Luis C. Pérez Tato (LCPT)"
__copyright__= "Copyright 2020, LCPT"
__license__= "GPL"
__version__= "3.0"
__email__= "l.pereztato@gmail.com"

import openpyxl

def to_str(s):
    try:
        return str(s)
    except ValueError:
        return ''
def to_int(s):
    try:
        return int(s)
    except ValueError:
        return None
def to_float(s):
    try:
        return float(s)
    except ValueError:
        return None

def name_columns_by_row(worksheet, row):
    ''' Return a dictionary containing the column indexes corresponding the
        contents of the given row.

    :param worksheet: worksheet to process.
    :param row: row containing the column names.
    '''
    retval= {}
    current= 0
    for col in worksheet.iter_cols(0, worksheet.max_column):
        value= col[row].value
        if(value):
            retval[value]= current
        current+= 1
    return retval

    
def import_load_combinations_from_sheet(openpyxlBook, sheetName, nameColumsRow, combNameColumn, actionLabels):
    ''' Import load combinations from an spreadsheet sheet

    :param openpyxlBook: spreadsheet book to import from.
    :param sheetName: name of the seet to import from.
    :param nameColumnsRow: row number of the column names.
    :param combNameColumn: label of the column that contains the 
                           combination name.
    :param actionLabels: labels of the actions.
    '''
    retval= dict()
    uls_sheet= openpyxlBook[sheetName]
    column_names= name_columns_by_row(uls_sheet, 1)
    rows= uls_sheet.rows
    rowIndex= 0
    for r in rows:
        if(rowIndex>=2):
            combName= to_str(r[column_names[combNameColumn]])
            combExpr= ''
            if(len(combName)>0):
                for a in actionLabels:
                    value= r[column_names[a]].value
                    if(value):
                        coef= to_float(value)
                        if(coef):
                            if(coef!= 0):
                                if(len(combExpr)>0):
                                    combExpr+= '+'
                                combExpr+= str(coef)+'*'+a
                    else:
                        break
                if(len(combExpr)>0):
                    retval[combName]= combExpr
        rowIndex+= 1
    return retval
    
def import_load_combinations_from_book(fileName, sheetNames, nameColumsRow, combNameColumn, actionLabels, dataOnly= False):
    ''' Import load combinations from an spreadsheet sheet

    :param fileName: spreadsheet file name import from.
    :param sheetNames: names of the seets to import from.
    :param nameColumnsRow: row number of the column names.
    :param combNameColumn: label of the column that contains the 
                           combination name.
    :param actionLabels: labels of the actions.
    :param dataOnly: read last computed values for cells that contain a formula.
    '''
    retval= dict()
    book= openpyxl.load_workbook(fileName, data_only= dataOnly)
    for name in sheetNames:
        retval[name]= import_load_combinations_from_sheet(book, name, nameColumsRow, combNameColumn, actionLabels)
    return retval
